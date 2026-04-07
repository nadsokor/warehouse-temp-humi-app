import re
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO

from openpyxl import load_workbook


def normalize_pandas_freq(freq: str) -> str:
    """
    pandas 2.2+ 不再把 T 当作「分钟」别名，旧写法 30T 会报错。
    将「数字+T」转为「数字+min」，其余原样返回。
    """
    s = (freq or "").strip()
    if not s:
        raise ValueError("时间网格频率不能为空")
    m = re.fullmatch(r"(\d+)T", s, flags=re.IGNORECASE)
    if m:
        return f"{m.group(1)}min"
    return s


def _normalize_header_label(name) -> str:
    """与 openpyxl 读到的表头使用同一套规则，便于对齐列号。"""
    if name is None:
        return ""
    s = str(name).strip()
    if s.startswith("\ufeff"):
        s = s.lstrip("\ufeff").strip()
    return s


def normalize_excel_columns(df: pd.DataFrame) -> pd.DataFrame:
    """去掉表头首尾空格与 BOM，避免与模板列名不一致。"""
    df = df.copy()
    df.columns = [_normalize_header_label(c) for c in df.columns]
    return df


def openpyxl_cell_to_id_string(cell) -> str:
    """
    从 openpyxl 单元格取出与 Excel 一致的展示语义：
    - 文本单元格：原样（含空格、制表符、前导零字符串等）。
    - 数字单元格：按存储值转字符串（整数无 .0）。
    """
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, bool):
        return str(v)
    dt = getattr(cell, "data_type", "")
    if dt == "e":
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, float):
        if np.isnan(v):
            return ""
        fv = float(v)
        if fv == int(fv):
            return str(int(fv))
        return format(fv, ".15g")
    if hasattr(v, "strftime"):
        return str(v)
    return str(v)


def overlay_id_columns_from_openpyxl(df: pd.DataFrame, file_bytes: bytes) -> pd.DataFrame:
    """
    用 openpyxl 按单元格重读「管理主机编号」「仪表编号」，避免 pandas 把文本当数字解析后丢失格式。
    失败时回退到 pandas 列 + normalize_excel_text_id_value。
    """
    df = df.copy()
    try:
        bio = BytesIO(file_bytes)
        wb = load_workbook(bio, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
            name_to_j = {}
            for j, c in enumerate(header_cells):
                key = _normalize_header_label(c.value)
                if key and key not in name_to_j:
                    name_to_j[key] = j
            n = len(df)
            data_rows = list(ws.iter_rows(min_row=2, max_row=n + 1, values_only=False))
            for col_name in ("管理主机编号", "仪表编号"):
                if col_name not in df.columns:
                    continue
                j = name_to_j.get(col_name)
                if j is None:
                    continue
                out = []
                for i in range(n):
                    if i >= len(data_rows):
                        out.append("")
                        continue
                    row = data_rows[i]
                    if j >= len(row):
                        out.append("")
                    else:
                        out.append(openpyxl_cell_to_id_string(row[j]))
                df[col_name] = out
        finally:
            wb.close()
    except Exception:
        for col in ("管理主机编号", "仪表编号"):
            if col in df.columns:
                df[col] = df[col].map(normalize_excel_text_id_value)
    return df


def normalize_excel_text_id_value(x) -> str:
    """
    无原始 xlsx 字节时的回退：尽量不改动文本；数字按语义转字符串（整数无 .0）。
    """
    if pd.isna(x):
        return ""
    # 文本：不做 strip，与输入完全一致
    if isinstance(x, str):
        return x
    if isinstance(x, np.str_):
        return str(x)
    if isinstance(x, (bool, np.bool_)):
        return str(x)
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        xf = float(x)
        if np.isnan(xf):
            return ""
        if xf == int(xf):
            return str(int(xf))
        return format(xf, ".15g")
    return str(x)


def process_dataframe(df, params, progress_callback=None, raw_excel_bytes=None):
    df = normalize_excel_columns(df)
    if raw_excel_bytes is not None:
        df = overlay_id_columns_from_openpyxl(df, raw_excel_bytes)
    else:
        for _col in ("管理主机编号", "仪表编号"):
            if _col in df.columns:
                df[_col] = df[_col].map(normalize_excel_text_id_value)
    # 文本类列不去首尾空白（含空格、制表符），与 Excel 单元格原文一致
    if "仪表名称" in df.columns:
        df["仪表名称"] = df["仪表名称"].apply(lambda v: "" if pd.isna(v) else str(v))
    freq = normalize_pandas_freq(params["freq"])
    temp_min = params["temp_min"]
    temp_target_min = params["temp_target_min"]
    temp_target_max = params["temp_target_max"]
    temp_max = params["temp_max"]
    humi_accept_min = params["humi_accept_min"]
    humi_accept_max = params["humi_accept_max"]
    humi_min = params["humi_min"]
    humi_max = params["humi_max"]
    temp_step_max = params["temp_step_max"]
    humi_step_max = params["humi_step_max"]
    smooth_max_diff_temp = params["smooth_max_diff_temp"]
    smooth_max_diff_humi = params["smooth_max_diff_humi"]
    random_seed = params["random_seed"]

    if random_seed is not None:
        np.random.seed(random_seed)

    original_columns = df.columns.tolist()

    required_cols = ["采集时间", "仪表名称", "仪表编号", "管理主机编号", "温度℃", "湿度%RH"]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        raise ValueError(
            f"缺少必要列: {missing_cols}。当前文件表头为: {list(df.columns)}"
        )

    df["采集时间"] = pd.to_datetime(df["采集时间"], errors="coerce")
    df["温度℃"] = pd.to_numeric(df["温度℃"], errors="coerce")
    df["湿度%RH"] = pd.to_numeric(df["湿度%RH"], errors="coerce")
    df["采集时间"] = df["采集时间"].dt.floor(freq)

    start_time = df["采集时间"].min().replace(hour=0, minute=0, second=0, microsecond=0)
    end_time = df["采集时间"].max().replace(hour=23, minute=30, second=0, microsecond=0)
    time_range = pd.date_range(start=start_time, end=end_time, freq=freq)

    def clip_round(val, vmin, vmax):
        return round(min(max(float(val), vmin), vmax), 1)

    def synthesize_from_last(last_row, t_point):
        new_row = last_row.copy()
        new_row["采集时间"] = t_point

        temp_diff = np.random.uniform(-temp_step_max, temp_step_max)
        humidity_diff = np.random.uniform(-humi_step_max, humi_step_max)

        last_temp = pd.to_numeric(last_row["温度℃"], errors="coerce")
        last_humi = pd.to_numeric(last_row["湿度%RH"], errors="coerce")
        if pd.isna(last_temp):
            last_temp = (temp_target_min + temp_target_max) / 2
        if pd.isna(last_humi):
            last_humi = (humi_min + humi_max) / 2

        new_row["温度℃"] = clip_round(last_temp + temp_diff, temp_target_min, temp_target_max)
        new_row["湿度%RH"] = clip_round(last_humi + humidity_diff, humi_min, humi_max)
        return new_row

    def should_keep_row(row):
        t = pd.to_numeric(row["温度℃"], errors="coerce")
        h = pd.to_numeric(row["湿度%RH"], errors="coerce")
        return (pd.notna(t) and temp_min <= float(t) <= temp_max) and (
            pd.notna(h) and humi_accept_min <= float(h) <= humi_accept_max
        )

    def process_group(group_df, full_time_index):
        group_df = (
            group_df.sort_values("采集时间")
            .drop_duplicates(subset=["采集时间"], keep="first")
            .reset_index(drop=True)
        )
        index_map = {t: i for i, t in enumerate(group_df["采集时间"])}

        result_rows = []
        have_prev = False
        last_row = None

        for tp in full_time_index:
            if tp in index_map:
                row = group_df.loc[index_map[tp]].copy()
                if should_keep_row(row):
                    humi_val = pd.to_numeric(row["湿度%RH"], errors="coerce")
                    if pd.notna(humi_val):
                        row["湿度%RH"] = clip_round(humi_val, humi_min, humi_max)
                    result_rows.append(row)
                    last_row = row
                    have_prev = True
                else:
                    if have_prev:
                        new_row = synthesize_from_last(last_row, tp)
                        result_rows.append(new_row)
                        last_row = new_row
            else:
                if have_prev:
                    new_row = synthesize_from_last(last_row, tp)
                    result_rows.append(new_row)
                    last_row = new_row

        if not result_rows:
            return pd.DataFrame(columns=group_df.columns)

        out = pd.DataFrame(result_rows)
        for col in original_columns:
            if col not in out.columns:
                out[col] = np.nan
        out = out[[c for c in original_columns if c in out.columns] + [c for c in out.columns if c not in original_columns]]
        return out

    def smooth_humidity(group):
        group = group.sort_values("采集时间").copy()
        if "湿度%RH" not in group.columns:
            return group
        vals = pd.to_numeric(group["湿度%RH"], errors="coerce").to_numpy(dtype="float64", copy=True)
        for i in range(1, len(vals)):
            if pd.isna(vals[i - 1]) or pd.isna(vals[i]):
                continue
            diff = vals[i] - vals[i - 1]
            if abs(diff) > smooth_max_diff_humi:
                adjust = np.random.uniform(0, smooth_max_diff_humi)
                vals[i] = round(vals[i - 1] + (adjust if diff > 0 else -adjust), 1)
                vals[i] = float(clip_round(vals[i], humi_min, humi_max))
        group["湿度%RH"] = vals
        return group

    def smooth_temp(group):
        group = group.sort_values("采集时间").copy()
        if "温度℃" not in group.columns:
            return group
        vals = pd.to_numeric(group["温度℃"], errors="coerce").to_numpy(dtype="float64", copy=True)
        for i in range(1, len(vals)):
            if pd.isna(vals[i - 1]) or pd.isna(vals[i]):
                continue
            diff = vals[i] - vals[i - 1]
            if abs(diff) > smooth_max_diff_temp:
                adjust = np.random.uniform(0, smooth_max_diff_temp)
                vals[i] = round(vals[i - 1] + (adjust if diff > 0 else -adjust), 1)
                vals[i] = float(clip_round(vals[i], temp_target_min, temp_target_max))
        group["温度℃"] = vals
        return group

    result_parts = []
    groups = list(df.groupby(["管理主机编号", "仪表编号"], dropna=False))
    total_groups = max(len(groups), 1)
    for idx, (_, g) in enumerate(groups, start=1):
        processed = process_group(g, time_range)
        if not processed.empty:
            result_parts.append(processed)
        if progress_callback is not None:
            progress_callback(idx / total_groups)

    result_df = pd.concat(result_parts, ignore_index=True) if result_parts else pd.DataFrame(columns=original_columns)
    result_df = result_df[result_df["采集时间"].isin(time_range)].copy()

    # 不用 groupby().apply()：新版 pandas 可能不把分组列放进子表，导致结果缺「管理主机编号」等列。
    if not result_df.empty:
        smooth_parts = []
        for _, g in result_df.groupby(["管理主机编号", "仪表编号"], dropna=False):
            g = smooth_humidity(g)
            g = smooth_temp(g)
            smooth_parts.append(g)
        result_df = pd.concat(smooth_parts, ignore_index=True)

    for col in original_columns:
        if col not in result_df.columns:
            result_df[col] = np.nan
    result_df = result_df[[c for c in original_columns if c in result_df.columns] + [c for c in result_df.columns if c not in original_columns]]

    result_df["采集时间"] = (
        pd.to_datetime(result_df["采集时间"], errors="coerce")
        .dt.strftime("%Y-%m-%d %H:%M:%S")
        .astype(str)
    )
    return result_df


def render_app():
    st.set_page_config(page_title="温湿度数据处理", layout="wide")
    st.title("温湿度数据处理")
    st.write("上传 Excel 后可按参数自动补全、裁剪与平滑温湿度数据，并导出新文件。")

    with st.sidebar:
        st.header("输入与参数")
        uploaded = st.file_uploader("上传 Excel 文件", type=["xlsx"])
        freq = st.text_input(
            "时间网格频率",
            value="30min",
            help="时间对齐的网格频率。半小时请填 30min（旧版脚本里的 30T 会自动换算）。",
        )

        st.subheader("温度范围")
        temp_min = st.number_input("温度接受下限", value=15.0, step=0.1, format="%.1f", help="用于判定是否保留原始温度的下限。")
        temp_target_min = st.number_input("温度生成下限", value=15.8, step=0.1, format="%.1f", help="生成/裁剪温度的下限。")
        temp_target_max = st.number_input("温度生成上限", value=18.5, step=0.1, format="%.1f", help="生成/裁剪温度的上限。")
        temp_max = st.number_input("温度接受上限", value=19.5, step=0.1, format="%.1f", help="用于判定是否保留原始温度的上限。")

        st.subheader("湿度范围")
        humi_accept_min = st.number_input("湿度接受下限", value=35.5, step=0.1, format="%.1f", help="用于判定是否保留原始湿度的下限。")
        humi_accept_max = st.number_input("湿度接受上限", value=74.5, step=0.1, format="%.1f", help="用于判定是否保留原始湿度的上限。")
        humi_min = st.number_input("湿度生成下限", value=36.5, step=0.1, format="%.1f", help="生成/裁剪湿度的下限。")
        humi_max = st.number_input("湿度生成上限", value=73.5, step=0.1, format="%.1f", help="生成/裁剪湿度的上限。")

        st.subheader("生成与平滑")
        temp_step_max = st.number_input("温度漂移幅度", value=0.5, step=0.1, format="%.1f", help="生成缺失温度时的随机漂移最大幅度。")
        humi_step_max = st.number_input("湿度漂移幅度", value=1.0, step=0.1, format="%.1f", help="生成缺失湿度时的随机漂移最大幅度。")
        smooth_max_diff_temp = st.number_input("温度平滑最大差值", value=1.0, step=0.1, format="%.1f", help="平滑阶段相邻温差最大值。")
        smooth_max_diff_humi = st.number_input("湿度平滑最大差值", value=1.0, step=0.1, format="%.1f", help="平滑阶段相邻湿差最大值。")

        seed_text = st.text_input("随机种子（可选）", value="", help="填写整数可让结果可复现，留空则随机。")
        random_seed = int(seed_text) if seed_text.strip().isdigit() else None

        output_name = st.text_input("导出文件名", value="处理后的数据.xlsx", help="下载时的文件名。")
        run_btn = st.button("开始处理")

    if not uploaded:
        st.info("请先上传 Excel 文件。")
        return

    if not run_btn:
        st.info("已上传文件，点击“开始处理”。")
        return

    params = {
        "freq": freq,
        "temp_min": temp_min,
        "temp_target_min": temp_target_min,
        "temp_target_max": temp_target_max,
        "temp_max": temp_max,
        "humi_accept_min": humi_accept_min,
        "humi_accept_max": humi_accept_max,
        "humi_min": humi_min,
        "humi_max": humi_max,
        "temp_step_max": temp_step_max,
        "humi_step_max": humi_step_max,
        "smooth_max_diff_temp": smooth_max_diff_temp,
        "smooth_max_diff_humi": smooth_max_diff_humi,
        "random_seed": random_seed,
    }

    try:
        raw = uploaded.getvalue()
        df = pd.read_excel(BytesIO(raw), engine="openpyxl")
        progress = st.progress(0, text="处理中：正在分组处理数据...")

        def on_progress(value):
            progress.progress(int(value * 70), text="处理中：正在分组处理数据...")

        result_df = process_dataframe(df, params, progress_callback=on_progress, raw_excel_bytes=raw)
        progress.progress(85, text="处理中：平滑与整理中...")
    except Exception as exc:
        st.error(f"处理失败：{exc}")
        return

    st.success(f"处理完成，生成 {len(result_df)} 行。")
    st.dataframe(result_df.head(50), use_container_width=True)

    progress.progress(90, text="处理中：生成导出文件...")
    with st.spinner("正在生成可下载文件，请稍候..."):
        output = BytesIO()
        result_df.to_excel(output, index=False, header=True)
        output.seek(0)
    progress.progress(100, text="处理完成，可下载。")
    st.download_button(
        label="下载处理后的数据",
        data=output,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    render_app()
